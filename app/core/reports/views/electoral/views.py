import json
from io import BytesIO

from django.conf import settings
from django.db.models import IntegerField, Max, Value
from django.db.models.aggregates import Count
from django.db.models.functions import Cast, Coalesce, Trim, Upper
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.template.loader import render_to_string
from django.urls import reverse_lazy
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import FormView, TemplateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from core.electoral.models import Elector, LocalVotacion, TipoVoto
from core.reports.forms import FormFilterGenerarPDFMesa, ReportForm
from core.reports.jasperbase import JasperReportBase
from core.security.mixins import ModuleMixin
from core.security.models import Module

try:
    from weasyprint import HTML
except ImportError:
    HTML = None


class ReporteMesaPreviewView(ModuleMixin, TemplateView):
    template_name = "electoral/reports/reporte_mesa.html"

    def _get_mesas_data(self):
        qs = (
            Elector.objects.filter(
                distrito=self.request.user.distrito,
                pasoxmv="S",
            )
            .annotate(mesa_int=Cast("mesa", IntegerField()))
            .values("local_votacion__denominacion", "mesa")
            .annotate(cantidad_votos=Count("id"))
            .order_by("local_votacion__denominacion", "mesa_int", "mesa")
        )

        mesas = []
        for row in qs:
            mesas.append(
                {
                    "local_votacion": row["local_votacion__denominacion"]
                    or "SIN LOCAL",
                    "numero_mesa": row["mesa"],
                    "pasoxmv": row["cantidad_votos"],
                }
            )

        total_votos = sum(int(item["pasoxmv"]) for item in mesas)
        return mesas, total_votos

    def post(self, request, *args, **kwargs):
        action = request.POST.get("action")
        if action == "sync_data":
            mesas, total_votos = self._get_mesas_data()
            return JsonResponse({"data": mesas, "total_votos": total_votos})
        return JsonResponse({"error": "Accion no valida"}, status=400)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        mesas, total_votos = self._get_mesas_data()

        context["nombre_local"] = ""
        context["distrito"] = getattr(self.request.user.distrito, "denominacion", "")
        context["title"] = "Reporte de Votos por Mesa"
        context["mesas"] = mesas
        context["total_votos"] = total_votos
        return context


class ReporteVotoTipoMesaView(ModuleMixin, TemplateView):
    template_name = "electoral/reports/reporte_tipo_voto_mesa.html"

    @staticmethod
    def _parse_int_values(request, keys):
        values = []
        for key in keys:
            values.extend(request.GET.getlist(key))
            values.extend(request.POST.getlist(key))

            raw = request.GET.get(key)
            if raw:
                values.extend(str(raw).split(","))

            raw = request.POST.get(key)
            if raw:
                values.extend(str(raw).split(","))

        parsed = []
        for value in values:
            text = str(value).strip()
            if not text:
                continue
            try:
                parsed.append(int(text))
            except (TypeError, ValueError):
                continue

        # Mantener orden y remover duplicados.
        return list(dict.fromkeys(parsed))

    @staticmethod
    def _parse_flag_value(request, keys):
        for key in keys:
            raw = request.GET.get(key)
            if raw is None:
                raw = request.POST.get(key)
            if raw is None:
                continue

            value = str(raw).strip().upper()
            if not value or value in {"ALL", "TODOS", "*"}:
                return None
            if value in {"S", "N"}:
                return value

        return None

    @staticmethod
    def _parse_int_values_from_post(request, keys):
        values = []
        for key in keys:
            values.extend(request.POST.getlist(key))

            raw = request.POST.get(key)
            if raw:
                values.extend(str(raw).split(","))

        parsed = []
        for value in values:
            text = str(value).strip()
            if not text:
                continue
            try:
                parsed.append(int(text))
            except (TypeError, ValueError):
                continue

        return list(dict.fromkeys(parsed))

    @staticmethod
    def _parse_flag_value_from_post(request, keys):
        for key in keys:
            raw = request.POST.get(key)
            if raw is None:
                continue

            value = str(raw).strip().upper()
            if not value or value in {"ALL", "TODOS", "*"}:
                return None
            if value in {"S", "N"}:
                return value

        return None

    def _get_elector_filters(self):
        # Evita que parametros GET de navegacion (querystring) apliquen filtros ocultos.
        local_ids = self._parse_int_values_from_post(
            self.request,
            ["local_votacion_id", "local_votacion_ids"],
        )
        tipo_voto_ids = self._parse_int_values_from_post(
            self.request,
            ["tipo_voto_id", "tipo_voto_ids"],
        )
        pasoxpc = self._parse_flag_value_from_post(self.request, ["pasoxpc"])
        pasoxmv = self._parse_flag_value_from_post(self.request, ["pasoxmv"])

        return {
            "local_ids": local_ids,
            "tipo_voto_ids": tipo_voto_ids,
            "pasoxpc": pasoxpc,
            "pasoxmv": pasoxmv,
        }

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        try:
            action = request.POST.get("action", "")

            if action == "refresh_stats":
                (
                    tipos_voto,
                    rows,
                    paso_local_totals,
                    paso_global_totals,
                    resumen_local_totals,
                    resumen_global_totals,
                ) = self._build_stats_data()
                return JsonResponse(
                    {
                        "tipos_voto": tipos_voto,
                        "rows_data": rows,
                        "paso_local_totals": paso_local_totals,
                        "paso_global_totals": paso_global_totals,
                        "resumen_local_totals": resumen_local_totals,
                        "resumen_global_totals": resumen_global_totals,
                    }
                )

            if action == "generate_excel_custom":
                payload_str = request.POST.get("datatable_payload", "")
                if not payload_str:
                    return JsonResponse(
                        {"error": "No se recibió el payload del DataTable"},
                        status=400,
                    )

                payload = json.loads(payload_str)
                return self._generate_excel_from_payload(payload)

            if action != "generate_pdf_weasy":
                return JsonResponse({"error": "Acción no válida"}, status=400)

            if HTML is None:
                return HttpResponse(
                    "Error: WeasyPrint no está instalado en el servidor.",
                    status=500,
                    content_type="text/plain",
                )

            from itertools import groupby
            from operator import itemgetter

            (
                tipos_voto,
                rows,
                _paso_local_totals,
                _paso_global_totals,
                _resumen_local_totals,
                _resumen_global_totals,
            ) = self._build_stats_data()

            grupos = []
            for local_name, group_iter in groupby(rows, key=itemgetter("local")):
                group_rows = list(group_iter)
                grupos.append(
                    {
                        "local": local_name,
                        "rows": group_rows,
                        "sub_tipo1": sum(r["tipo_1"] for r in group_rows),
                        "sub_otros": sum(r["total_otros"] for r in group_rows),
                        "sub_dif": sum(r["diferencia"] for r in group_rows),
                    }
                )

            context = {
                "distrito": getattr(request.user.distrito, "denominacion", ""),
                "tipos_voto": tipos_voto,
                "grupos": grupos,
                "total_tipo1": sum(g["sub_tipo1"] for g in grupos),
                "total_otros": sum(g["sub_otros"] for g in grupos),
                "total_dif": sum(g["sub_dif"] for g in grupos),
            }

            html_str = render_to_string(
                "electoral/reports/reporte_tipo_voto_mesa_pdf.html",
                context,
                request=request,
            )

            pdf_bytes = HTML(
                string=html_str,
                base_url=request.build_absolute_uri("/"),
            ).write_pdf()

            response = HttpResponse(pdf_bytes, content_type="application/pdf")
            response["Content-Disposition"] = (
                'inline; filename="estadistica_votos_tipo.pdf"'
            )
            return response

        except Exception as exc:
            return HttpResponse(
                f"Error generando archivo: {type(exc).__name__}: {exc}",
                status=500,
                content_type="text/plain; charset=utf-8",
            )

    def _generate_excel_from_payload(self, payload):
        headers = payload.get("headers", {})
        groups = payload.get("groups", [])
        totals = payload.get("totals", {})

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"

        title = payload.get("title") or "Estadistica de Votos por Tipo"
        district = payload.get("district") or ""

        ws.merge_cells("A1:D1")
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:D2")
        ws["A2"] = f"Distrito: {district}"
        ws["A2"].alignment = Alignment(horizontal="center")

        row_idx = 4
        head_fill = PatternFill(
            start_color="E9ECEF", end_color="E9ECEF", fill_type="solid"
        )
        sub_fill = PatternFill(
            start_color="F8F9FA", end_color="F8F9FA", fill_type="solid"
        )
        total_fill = PatternFill(
            start_color="E2E8F0", end_color="E2E8F0", fill_type="solid"
        )

        ws.cell(row=row_idx, column=1, value=headers.get("mesa", "Mesa"))
        ws.cell(row=row_idx, column=2, value=headers.get("tipo_1", "Tipo 1"))
        ws.cell(row=row_idx, column=3, value=headers.get("total_otros", "Otros tipos"))
        ws.cell(row=row_idx, column=4, value=headers.get("diferencia", "Diferencia"))
        for col in range(1, 5):
            ws.cell(row=row_idx, column=col).font = Font(bold=True)
            ws.cell(row=row_idx, column=col).fill = head_fill
        row_idx += 2

        for group in groups:
            local_name = group.get("local", "SIN LOCAL")
            ws.merge_cells(
                start_row=row_idx, start_column=1, end_row=row_idx, end_column=4
            )
            ws.cell(row=row_idx, column=1, value=f"LOCAL DE VOTACION: {local_name}")
            ws.cell(row=row_idx, column=1).font = Font(bold=True)
            ws.cell(row=row_idx, column=1).fill = head_fill
            row_idx += 1

            for item in group.get("rows", []):
                ws.cell(row=row_idx, column=1, value=item.get("mesa", ""))
                ws.cell(row=row_idx, column=2, value=int(item.get("tipo_1", 0)))
                ws.cell(row=row_idx, column=3, value=int(item.get("total_otros", 0)))
                ws.cell(row=row_idx, column=4, value=int(item.get("diferencia", 0)))
                row_idx += 1

            ws.cell(row=row_idx, column=1, value=f"Subtotal {local_name}")
            ws.cell(row=row_idx, column=2, value=int(group.get("subtotal_tipo1", 0)))
            ws.cell(row=row_idx, column=3, value=int(group.get("subtotal_otros", 0)))
            ws.cell(
                row=row_idx, column=4, value=int(group.get("subtotal_diferencia", 0))
            )
            for col in range(1, 5):
                ws.cell(row=row_idx, column=col).font = Font(bold=True)
                ws.cell(row=row_idx, column=col).fill = sub_fill
            row_idx += 2

        ws.cell(row=row_idx, column=1, value="TOTAL GENERAL")
        ws.cell(row=row_idx, column=2, value=int(totals.get("tipo_1", 0)))
        ws.cell(row=row_idx, column=3, value=int(totals.get("total_otros", 0)))
        ws.cell(row=row_idx, column=4, value=int(totals.get("diferencia", 0)))
        for col in range(1, 5):
            ws.cell(row=row_idx, column=col).font = Font(bold=True)
            ws.cell(row=row_idx, column=col).fill = total_fill

        ws.column_dimensions["A"].width = 36
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18

        for row in ws.iter_rows(min_row=1, max_row=row_idx, min_col=1, max_col=4):
            for cell in row:
                if cell.column == 1:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="right", vertical="center")

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = (
            'attachment; filename="estadistica_votos_tipo.xlsx"'
        )
        return response

    def _build_stats_data(self):
        def normalize_local_name(value):
            local_name = str(value or "SIN LOCAL").strip()
            return local_name or "SIN LOCAL"

        def normalize_local_id(value):
            try:
                return int(value)
            except (TypeError, ValueError):
                return 0

        filtros = self._get_elector_filters()

        all_qs = Elector.objects.filter(
            distrito=self.request.user.distrito,
        )

        # QS base de la tabla: puede filtrar por local y tipo_voto.
        table_qs = all_qs

        if filtros["local_ids"]:
            table_qs = table_qs.filter(local_votacion_id__in=filtros["local_ids"])

        if filtros["tipo_voto_ids"]:
            table_qs = table_qs.filter(tipo_voto_id__in=filtros["tipo_voto_ids"])

        normalized_table_qs = table_qs.annotate(
            pasoxmv_norm=Coalesce(Upper(Trim("pasoxmv")), Value("N")),
            pasoxpc_norm=Coalesce(Upper(Trim("pasoxpc")), Value("N")),
        )

        # QS base del RESUMEN PASO: independiente de filtros de tabla (tipo/local request).
        normalized_paso_qs = all_qs.annotate(
            pasoxmv_norm=Coalesce(Upper(Trim("pasoxmv")), Value("N")),
            pasoxpc_norm=Coalesce(Upper(Trim("pasoxpc")), Value("N")),
        )

        pasoxmv_filter = filtros["pasoxmv"] if filtros["pasoxmv"] in {"S", "N"} else "S"
        pasoxpc_filter = filtros["pasoxpc"] if filtros["pasoxpc"] in {"S", "N"} else "S"

        # Base de grilla: mantiene criterio de PASOXMV sin combinar con PASOXPC.
        base_qs = normalized_table_qs.filter(pasoxmv_norm=pasoxmv_filter)

        # RESUMEN PASO para la tabla: misma base que la grilla para evitar desfasajes.
        pasoxmv_table_qs = base_qs
        pasoxpc_table_qs = normalized_table_qs.filter(pasoxpc_norm=pasoxpc_filter)

        # Mantener tambien los globales independientes para otros usos.
        pasoxmv_summary_qs = normalized_paso_qs.filter(pasoxmv_norm=pasoxmv_filter)
        pasoxpc_summary_qs = normalized_paso_qs.filter(pasoxpc_norm=pasoxpc_filter)

        tipo_ids = list(
            base_qs.exclude(tipo_voto_id__isnull=True)
            .values_list("tipo_voto_id", flat=True)
            .distinct()
        )

        tipos_voto = list(
            TipoVoto.objects.filter(id__in=tipo_ids)
            .order_by("id", "denominacion")
            .values("id", "cod", "denominacion")
        )

        mesas_base = list(
            base_qs.values("local_votacion_id", "local_votacion__denominacion", "mesa")
            .annotate(mesa_int=Cast("mesa", IntegerField()))
            .order_by("local_votacion__denominacion", "mesa_int", "mesa")
            .values("local_votacion_id", "local_votacion__denominacion", "mesa")
            .distinct()
        )

        conteos_qs = (
            base_qs.exclude(tipo_voto_id__isnull=True)
            .annotate(mesa_int=Cast("mesa", IntegerField()))
            .values(
                "local_votacion_id",
                "local_votacion__denominacion",
                "mesa",
                "tipo_voto_id",
            )
            .annotate(cantidad=Count("id"))
            .order_by(
                "local_votacion__denominacion", "mesa_int", "mesa", "tipo_voto_id"
            )
        )

        conteos_null_qs = (
            base_qs.filter(tipo_voto_id__isnull=True)
            .annotate(mesa_int=Cast("mesa", IntegerField()))
            .values(
                "local_votacion_id",
                "local_votacion__denominacion",
                "mesa",
            )
            .annotate(cantidad_null=Count("id"))
            .order_by("local_votacion__denominacion", "mesa_int", "mesa")
        )

        pasoxmv_counts_map = {
            f"{normalize_local_id(row['local_votacion_id'])}|{(row['mesa'] or '')}": int(
                row["pasoxmv_count"]
            )
            for row in pasoxmv_table_qs.values(
                "local_votacion_id", "local_votacion__denominacion", "mesa"
            ).annotate(pasoxmv_count=Count("id"))
        }

        pasoxpc_counts_map = {
            f"{normalize_local_id(row['local_votacion_id'])}|{(row['mesa'] or '')}": int(
                row["pasoxpc_count"]
            )
            for row in pasoxpc_table_qs.values(
                "local_votacion_id", "local_votacion__denominacion", "mesa"
            ).annotate(pasoxpc_count=Count("id"))
        }

        pasoxmv_local_totals = {
            str(normalize_local_id(row["local_votacion_id"])): int(row["pasoxmv_count"])
            for row in pasoxmv_table_qs.values(
                "local_votacion_id", "local_votacion__denominacion"
            ).annotate(pasoxmv_count=Count("id"))
        }

        pasoxpc_local_totals = {
            str(normalize_local_id(row["local_votacion_id"])): int(row["pasoxpc_count"])
            for row in pasoxpc_table_qs.values(
                "local_votacion_id", "local_votacion__denominacion"
            ).annotate(pasoxpc_count=Count("id"))
        }

        paso_local_totals = {}
        for local_name in set(pasoxmv_local_totals.keys()) | set(
            pasoxpc_local_totals.keys()
        ):
            paso_local_totals[local_name] = {
                "pasoxmv": int(pasoxmv_local_totals.get(local_name, 0)),
                "pasoxpc": int(pasoxpc_local_totals.get(local_name, 0)),
            }

        paso_global_totals = {
            "pasoxmv": int(pasoxmv_table_qs.count()),
            "pasoxpc": int(pasoxpc_table_qs.count()),
        }

        resumen_tipo1_local_totals = {
            str(normalize_local_id(row["local_votacion_id"])): int(row["cantidad"])
            for row in base_qs.filter(tipo_voto_id=1)
            .values("local_votacion_id")
            .annotate(cantidad=Count("id"))
        }

        resumen_tipo2_local_totals = {
            str(normalize_local_id(row["local_votacion_id"])): int(row["cantidad"])
            for row in base_qs.filter(tipo_voto_id=2)
            .values("local_votacion_id")
            .annotate(cantidad=Count("id"))
        }

        resumen_otros_local_totals = {
            str(normalize_local_id(row["local_votacion_id"])): int(row["cantidad"])
            for row in base_qs.exclude(tipo_voto_id__in=[1, 2])
            .values("local_votacion_id")
            .annotate(cantidad=Count("id"))
        }

        resumen_local_totals = {}
        for local_id in (
            set(resumen_tipo1_local_totals.keys())
            | set(resumen_tipo2_local_totals.keys())
            | set(resumen_otros_local_totals.keys())
        ):
            total_v1 = int(resumen_tipo1_local_totals.get(local_id, 0))
            total_v2 = int(resumen_tipo2_local_totals.get(local_id, 0))
            total_otros = int(resumen_otros_local_totals.get(local_id, 0))
            resumen_local_totals[local_id] = {
                "total_v1": total_v1,
                "total_v2": total_v2,
                "total_otros": total_otros,
                "diferencia": total_v1 - (total_v2 + total_otros),
            }

        resumen_global_totals = {
            "total_v1": int(base_qs.filter(tipo_voto_id=1).count()),
            "total_v2": int(base_qs.filter(tipo_voto_id=2).count()),
            "total_otros": int(base_qs.exclude(tipo_voto_id__in=[1, 2]).count()),
        }
        resumen_global_totals["diferencia"] = resumen_global_totals["total_v1"] - (
            resumen_global_totals["total_v2"] + resumen_global_totals["total_otros"]
        )

        rows_map = {}
        for mesa_row in mesas_base:
            local_id = normalize_local_id(mesa_row["local_votacion_id"])
            local_name = normalize_local_name(mesa_row["local_votacion__denominacion"])
            mesa_num = mesa_row["mesa"] or ""
            key = f"{local_id}|{mesa_num}"
            rows_map[key] = {
                "local_id": local_id,
                "local": local_name,
                "mesa": mesa_num,
                "pasoxmv_count": int(pasoxmv_counts_map.get(key, 0)),
                "pasoxpc_count": int(pasoxpc_counts_map.get(key, 0)),
                "tipo_0": 0,
                **{f"tipo_{tipo['id']}": 0 for tipo in tipos_voto},
            }

        for count_row in conteos_qs:
            local_id = normalize_local_id(count_row["local_votacion_id"])
            local_name = normalize_local_name(count_row["local_votacion__denominacion"])
            mesa_num = count_row["mesa"] or ""
            key = f"{local_id}|{mesa_num}"
            if key not in rows_map:
                rows_map[key] = {
                    "local_id": local_id,
                    "local": local_name,
                    "mesa": mesa_num,
                    "pasoxmv_count": int(pasoxmv_counts_map.get(key, 0)),
                    "pasoxpc_count": int(pasoxpc_counts_map.get(key, 0)),
                    "tipo_0": 0,
                    **{f"tipo_{tipo['id']}": 0 for tipo in tipos_voto},
                }
            rows_map[key][f"tipo_{count_row['tipo_voto_id']}"] = count_row["cantidad"]

        for count_row in conteos_null_qs:
            local_id = normalize_local_id(count_row["local_votacion_id"])
            local_name = normalize_local_name(count_row["local_votacion__denominacion"])
            mesa_num = count_row["mesa"] or ""
            key = f"{local_id}|{mesa_num}"
            if key not in rows_map:
                rows_map[key] = {
                    "local_id": local_id,
                    "local": local_name,
                    "mesa": mesa_num,
                    "pasoxmv_count": int(pasoxmv_counts_map.get(key, 0)),
                    "pasoxpc_count": int(pasoxpc_counts_map.get(key, 0)),
                    "tipo_0": 0,
                    **{f"tipo_{tipo['id']}": 0 for tipo in tipos_voto},
                }
            rows_map[key]["tipo_0"] = int(count_row["cantidad_null"])

        tipo_principal_key = "tipo_1"
        for row in rows_map.values():
            base = row.get(tipo_principal_key, 0)
            otros = sum(
                v
                for k, v in row.items()
                if k.startswith("tipo_") and k != tipo_principal_key
            )
            row["total_otros"] = otros
            row["diferencia"] = base - otros

        rows = list(rows_map.values())
        rows.sort(
            key=lambda r: (
                r["local"],
                int(r["mesa"]) if str(r["mesa"]).isdigit() else 999999,
                str(r["mesa"]),
            )
        )

        return (
            tipos_voto,
            rows,
            paso_local_totals,
            paso_global_totals,
            resumen_local_totals,
            resumen_global_totals,
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        (
            tipos_voto,
            rows,
            paso_local_totals,
            paso_global_totals,
            resumen_local_totals,
            resumen_global_totals,
        ) = self._build_stats_data()

        context["title"] = "Estadistica de Votos por Tipo"
        context["distrito"] = getattr(self.request.user.distrito, "denominacion", "")
        context["tipos_voto"] = tipos_voto
        context["rows_data"] = rows
        context["tipos_voto_json"] = json.dumps(tipos_voto, ensure_ascii=False)
        context["rows_data_json"] = json.dumps(rows, ensure_ascii=False)
        context["paso_local_totals_json"] = json.dumps(
            paso_local_totals, ensure_ascii=False
        )
        context["paso_global_totals_json"] = json.dumps(
            paso_global_totals, ensure_ascii=False
        )
        context["resumen_local_totals_json"] = json.dumps(
            resumen_local_totals, ensure_ascii=False
        )
        context["resumen_global_totals_json"] = json.dumps(
            resumen_global_totals, ensure_ascii=False
        )
        context["list_url"] = reverse_lazy("reporte_tipo_voto_mesa")
        return context


"""Reporte de Barrios y Manzanas con Codigo"""


# class RptElectoral000ReportView(ModuleMixin, FormView):
#     template_name = "electoral/reports/rpt_electoral000.html"
#     form_class = ReportForm

#     @method_decorator(csrf_exempt)
#     def dispatch(self, request, *args, **kwargs):
#         return super().dispatch(request, *args, **kwargs)

#     def post(self, request, *args, **kwargs):
#         action = request.POST["action"]
#         data = {}
#         try:
#             if action == "search_report":
#                 data = []
#                 print(request.POST)
#                 seccional = (
#                     request.POST.getlist("seccional[]")
#                     if "seccional[]" in request.POST
#                     else None
#                 )
#                 seccional = seccional if seccional != [""] else None
#                 barrio = (
#                     request.POST.getlist("barrio[]")
#                     if "barrio[]" in request.POST
#                     else None
#                 )
#                 barrio = barrio if barrio != [""] else None
#                 manzana = (
#                     request.POST.getlist("manzana[]")
#                     if "manzana[]" in request.POST
#                     else None
#                 )
#                 manzana = manzana if manzana != [""] else None
#                 # end_date = request.POST['end_date']
#                 _where = "1=1"

#                 if seccional:
#                     _where += f" AND electoral_elector.seccional_id IN {seccional}"
#                 if barrio:
#                     _where += f" AND electoral_elector.barrio_id IN {barrio}"
#                 if manzana:
#                     _where += f" AND electoral_elector.manzana_id IN {manzana}"
#                 _where = _where.replace("[", "(").replace("]", ")")
#                 print(_where)
#                 qs = (
#                     Elector.objects.values(
#                         "barrio__id",
#                         "barrio__denominacion",
#                         "manzana__cod",
#                         "manzana__denominacion",
#                     )
#                     .filter(distrito=self.request.user.distrito)
#                     .extra(
#                         select={
#                             "barrio__cod": "CAST (electoral_elector.barrio_id AS INTEGER)"
#                         }
#                     )
#                     .annotate(cant_elector=Count(True))
#                     .extra(where=[_where])
#                     .order_by("barrio__cod", "manzana__cod")
#                 )
#                 for i in qs:
#                     item = {
#                         "barrio": f"({i['barrio__id']}) - {i['barrio__denominacion']}",
#                         "manzana": f"({i['barrio__id']} / {i['manzana__cod']}) - {i['manzana__denominacion']}",
#                         "cant_elector": i["cant_elector"],
#                     }
#                     data.append(item)
#                 # print(data)
#             else:
#                 data["error"] = "No ha ingresado una opción"
#         except Exception as e:
#             data["error"] = str(e)
#         return HttpResponse(json.dumps(data), content_type="application/json")

#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         context["form"] = ReportForm(usuario=self.request.user)
#         context["title"] = "Reporte de Barrios y Manzanas"
#         return context


class RptElectoral000ReportView(ModuleMixin, FormView):
    template_name = "electoral/reports/rpt_electoral000.html"
    form_class = ReportForm

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        action = request.POST.get("action")
        data = {}

        try:
            # --- ACCIÓN 1: BÚSQUEDA TRADICIONAL (CORREGIDA SIN SQL INJECTION) ---
            if action == "search_report":
                data = []

                # Extraer listas limpias
                seccional = [x for x in request.POST.getlist("seccional[]") if x != ""]
                barrio = [x for x in request.POST.getlist("barrio[]") if x != ""]
                manzana = [x for x in request.POST.getlist("manzana[]") if x != ""]

                # Construcción segura del QuerySet usando el ORM nativo de Django
                filters = {"distrito": self.request.user.distrito}
                if seccional:
                    filters["seccional_id__in"] = seccional
                if barrio:
                    filters["barrio_id__in"] = barrio
                if manzana:
                    filters["manzana_id__in"] = manzana

                qs = (
                    Elector.objects.filter(**filters)
                    .values(
                        "barrio__id",
                        "barrio__denominacion",
                        "manzana__cod",
                        "manzana__denominacion",
                        "barrio_id",  # Reemplaza el CAST manual usando el ID directo de la relación
                    )
                    .annotate(
                        cant_elector=Count("id")
                    )  # Count sobre la PK para mayor precisión
                    .order_by("barrio_id", "manzana__cod")
                )

                for i in qs:
                    item = {
                        "barrio": f"({i['barrio__id']}) - {i['barrio__denominacion']}",
                        "manzana": f"({i['barrio__id']} / {i['manzana__cod']}) - {i['manzana__denominacion']}",
                        "cant_elector": i["cant_elector"],
                    }
                    data.append(item)
                return HttpResponse(json.dumps(data), content_type="application/json")

            # --- NUEVA ACCIÓN 2: GENERAR EL REPORTE PDF CON WEASYPRINT ---
            elif action == "generate_pdf":
                # Capturamos los campos del formulario
                local_id = request.POST.get("local_votacion_id")
                mesa_num = request.POST.get("mesa")

                if not local_id:
                    return HttpResponse(
                        "Error: El ID del Local de Votación es requerido.", status=400
                    )

                # Determinamos la lista de mesas (Específica o todas las del local)
                if mesa_num and mesa_num.strip():
                    mesas_lista = [int(mesa_num)]
                else:
                    mesas_lista = (
                        Elector.objects.filter(local_votacion_id=local_id)
                        .values_list("mesa", flat=True)
                        .distinct()
                        .order_by("mesa")
                    )

                mesas_data = []
                columnas = 20

                # Armamos las matrices dinámicas por cada mesa encontrada
                for m in mesas_lista:
                    datos_mesa = Elector.objects.filter(
                        local_votacion_id=local_id, mesa=m
                    ).aggregate(max_orden=Max("orden"))

                    max_orden = datos_mesa["max_orden"] or 0

                    matriz_ordenes = []
                    fila_actual = []

                    for i in range(1, max_orden + 1):
                        fila_actual.append(i)
                        if len(fila_actual) == columnas:
                            matriz_ordenes.append(fila_actual)
                            fila_actual = []

                    if fila_actual:
                        while len(fila_actual) < columnas:
                            fila_actual.append(None)
                        matriz_ordenes.append(fila_actual)

                    # Ajuste dinámico de escala para forzar una sola hoja por mesa
                    total_filas = len(matriz_ordenes)
                    if total_filas > 15:
                        escala_css = "escala-micro"
                    elif total_filas > 10:
                        escala_css = "escala-chica"
                    else:
                        escala_css = "escala-normal"

                    mesas_data.append(
                        {
                            "numero_mesa": m,
                            "matriz": matriz_ordenes,
                            "max_orden": max_orden,
                            "escala_css": escala_css,
                        }
                    )

                # Datos del Local para el Encabezado
                nombre_local = f"LOCAL DE VOTACIÓN N° {local_id}"
                total_votos = sum(
                    int(mesa.get("max_orden") or 0) for mesa in mesas_data
                )
                context = {
                    "nombre_local": nombre_local,
                    "mesas": mesas_data,
                    "total_votos": total_votos,
                }

                # Renderizamos el template HTML a un String crudo
                html_string = render(
                    request, "electoral/reports/reporte_mesa.html", context
                ).content.decode("utf-8")

                if HTML is None:
                    return HttpResponse(
                        "Error: WeasyPrint no está configurado en el servidor.",
                        status=500,
                    )

                # Compilación y generación binaria del PDF
                html_pdf = HTML(
                    string=html_string, base_url=request.build_absolute_uri()
                )
                pdf_file = html_pdf.write_pdf()

                filename = (
                    f"Planilla_Local_{local_id}.pdf"
                    if not mesa_num
                    else f"Planilla_Mesa_{mesa_num}.pdf"
                )
                response = HttpResponse(pdf_file, content_type="application/pdf")
                response["Content-Disposition"] = f'inline; filename="{filename}"'
                return response

            else:
                data["error"] = "No ha ingresado una opción válida"

        except Exception as e:
            # Si el request es de tipo reporte JSON devolvemos el error estructurado,
            # si falla la acción del PDF, mostramos el mensaje directo.
            if action == "search_report":
                data["error"] = str(e)
                return HttpResponse(json.dumps(data), content_type="application/json")
            return HttpResponse(f"Error procesando reporte: {str(e)}", status=500)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = ReportForm(usuario=self.request.user)
        context["title"] = "Reporte de Barrios y Manzanas"
        return context


class RptPadron001ReportView(ModuleMixin, FormView):
    template_name = "electoral/reports/rpt_padron001.html"
    form_class = ReportForm

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        action = request.POST["action"]
        data = {}
        # print(request.POST)
        try:
            if action == "report":
                data = []
                tipo = request.POST["tipo"]
                local_votacion = (
                    request.POST.getlist("local_votacion")
                    if "local_votacion" in request.POST
                    else None
                )
                barrio = (
                    request.POST.getlist("barrio") if "barrio" in request.POST else None
                )
                tipo_voto = (
                    request.POST.getlist("tipo_voto")
                    if "tipo_voto" in request.POST
                    else None
                )
                # Tipo de Voto I - INDECISO es igual a NO DEFINIDOS null ver query reporte
                # CONFIG
                report = JasperReportBase()
                report.report_name = "rpt_padron001"
                report.report_url = reverse_lazy(report.report_name)
                report.report_title = (
                    Module.objects.filter(url=report.report_url).first().name
                )
                # PARAMETROS
                report.params["P_LOCAL_VOTACION_ID"] = (
                    ",".join(local_votacion) if local_votacion != [""] else None
                )
                report.params["P_BARRIO_ID"] = (
                    ",".join(barrio) if barrio != [""] else None
                )
                report.params["P_TIPO_VOTO_ID"] = (
                    ",".join(tipo_voto) if tipo_voto != [""] else None
                )

                return report.render_to_response(tipo)

            else:
                data["error"] = "No ha ingresado una opción"
        except Exception as e:
            data["error"] = str(e)
        return HttpResponse(json.dumps(data), content_type="application/json")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = ReportForm(usuario=self.request.user)
        context["title"] = "Reporte de Padron"
        context["action"] = "report"
        return context


"""Electores por Barrios y Manzanas"""


class RptElectoral001ReportView(ModuleMixin, FormView):
    template_name = "electoral/reports/rpt_electoral001.html"
    form_class = ReportForm

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        action = request.POST["action"]
        data = {}
        # print(request.POST)
        try:
            if action == "report":
                data = []
                tipo = request.POST["tipo"]
                local_votacion = (
                    request.POST.getlist("local_votacion")
                    if "local_votacion" in request.POST
                    else None
                )
                barrio = (
                    request.POST.getlist("barrio") if "barrio" in request.POST else None
                )
                manzana = (
                    request.POST.getlist("manzana")
                    if "manzana" in request.POST
                    else None
                )
                salto_pagina = (
                    request.POST.getlist("salto_pagina")
                    if "salto_pagina" in request.POST
                    else None
                )
                titulo_extra = (
                    request.POST.getlist("titulo_extra")
                    if "titulo_extra" in request.POST
                    else ""
                )
                # CONFIG
                report = JasperReportBase()
                report.report_name = "rpt_electoral001"

                report.report_url = reverse_lazy(report.report_name)
                report.report_title = (
                    Module.objects.filter(url=report.report_url).first().name
                )
                if len(titulo_extra):
                    report.report_title = titulo_extra[0]
                # PARAMETROS
                report.params["P_LOCAL_VOTACION_ID"] = (
                    ",".join(local_votacion) if local_votacion != [""] else None
                )
                report.params["P_BARRIO_ID"] = (
                    ",".join(barrio) if barrio != [""] else None
                )
                report.params["P_MANZANA_ID"] = (
                    ",".join(manzana) if manzana != [""] else None
                )

                if not salto_pagina:
                    report.report_name = "rpt_electoral001_ss"

                return report.render_to_response(tipo)

            else:
                data["error"] = "No ha ingresado una opción"
        except Exception as e:
            data["error"] = str(e)
        return HttpResponse(json.dumps(data), content_type="application/json")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = ReportForm(usuario=self.request.user)
        context["title"] = "Reporte de Elector por Barrios y Manzanas"
        context["action"] = "report"
        return context


"""Electores por Barrios y Manzanas"""


class RptElectoral002ReportView(ModuleMixin, FormView):
    template_name = "electoral/reports/rpt_electoral002.html"
    form_class = ReportForm

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        action = request.POST["action"]
        data = {}
        # print(request.POST)
        try:
            if action == "report":
                data = []
                tipo = request.POST["tipo"]
                local_votacion = (
                    request.POST.getlist("local_votacion")
                    if "local_votacion" in request.POST
                    else None
                )
                barrio = (
                    request.POST.getlist("barrio") if "barrio" in request.POST else None
                )
                manzana = (
                    request.POST.getlist("manzana")
                    if "manzana" in request.POST
                    else None
                )
                salto_pagina = (
                    request.POST.getlist("salto_pagina")
                    if "salto_pagina" in request.POST
                    else None
                )
                titulo_extra = (
                    request.POST.getlist("titulo_extra")
                    if "titulo_extra" in request.POST
                    else ""
                )
                # CONFIG
                report = JasperReportBase()
                report.report_name = "rpt_electoral002"
                report.report_url = reverse_lazy(report.report_name)

                report.report_title = (
                    Module.objects.filter(url=report.report_url).first().name
                )
                if len(titulo_extra):
                    report.report_title = titulo_extra[0]

                # PARAMETROS
                report.params["P_LOCAL_VOTACION_ID"] = (
                    ",".join(local_votacion) if local_votacion != [""] else None
                )
                report.params["P_BARRIO_ID"] = (
                    ",".join(barrio) if barrio != [""] else None
                )
                report.params["P_MANZANA_ID"] = (
                    ",".join(manzana) if manzana != [""] else None
                )

                if not salto_pagina:
                    report.report_name = "rpt_electoral002_ss"

                return report.render_to_response(tipo)

            else:
                data["error"] = "No ha ingresado una opción"
        except Exception as e:
            data["error"] = str(e)
        return HttpResponse(json.dumps(data), content_type="application/json")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = ReportForm(usuario=self.request.user)
        context["title"] = "Reporte de Elector por Barrios y Manzanas"
        context["action"] = "report"
        return context


"""Electores por Operadores"""


class RptElectoral003ReportView(ModuleMixin, FormView):
    template_name = "electoral/reports/rpt_electoral003.html"
    form_class = ReportForm

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        action = request.POST["action"]
        data = {}
        # print(request.POST)
        try:
            if action == "report":
                data = []
                tipo = request.POST["tipo"]
                local_votacion = (
                    request.POST.getlist("local_votacion")
                    if "local_votacion" in request.POST
                    else None
                )
                barrio = (
                    request.POST.getlist("barrio") if "barrio" in request.POST else None
                )
                manzana = (
                    request.POST.getlist("manzana")
                    if "manzana" in request.POST
                    else None
                )
                operador = (
                    request.POST.getlist("operador")
                    if "operador" in request.POST
                    else None
                )
                salto_pagina = (
                    request.POST.getlist("salto_pagina")
                    if "salto_pagina" in request.POST
                    else None
                )
                titulo_extra = (
                    request.POST.getlist("titulo_extra")
                    if "titulo_extra" in request.POST
                    else ""
                )
                # CONFIG
                report = JasperReportBase()
                report.report_name = "rpt_electoral003"
                report.report_url = reverse_lazy(report.report_name)

                report.report_title = (
                    Module.objects.filter(url=report.report_url).first().name
                )
                if len(titulo_extra):
                    report.report_title = titulo_extra[0]

                # PARAMETROS
                report.params["P_LOCAL_VOTACION_ID"] = (
                    ",".join(local_votacion) if local_votacion != [""] else None
                )
                report.params["P_BARRIO_ID"] = (
                    ",".join(barrio) if barrio != [""] else None
                )
                report.params["P_MANZANA_ID"] = (
                    ",".join(manzana) if manzana != [""] else None
                )
                report.params["P_OPERADOR_ID"] = (
                    ",".join(operador) if operador != [""] else None
                )

                if not salto_pagina:
                    report.report_name = "rpt_electoral003"

                return report.render_to_response(tipo)

            else:
                data["error"] = "No ha ingresado una opción"
        except Exception as e:
            data["error"] = str(e)
        return HttpResponse(json.dumps(data), content_type="application/json")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = ReportForm(usuario=self.request.user)
        context["title"] = "Reporte de Elector por Barrios y Manzanas"
        context["action"] = "report"
        return context


"""Estadistica de Votos Positivos vs Negativos"""


class RptEstadistica001ReportView(ModuleMixin, FormView):
    template_name = "electoral/reports/rpt_estadistica001.html"
    form_class = ReportForm

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        action = request.POST["action"]
        data = {}
        try:
            if action == "report":
                data = []
                tipo = request.POST["tipo"]
                local_votacion = (
                    request.POST.getlist("local_votacion")
                    if "local_votacion" in request.POST
                    else None
                )

                # CONFIG
                report = JasperReportBase()
                report.report_name = "rpt_estadistica001"
                report.report_url = reverse_lazy(report.report_name)
                report.report_title = (
                    Module.objects.filter(url=report.report_url).first().name
                )
                # PARAMETROS
                report.params["P_LOCAL_VOTACION_ID"] = (
                    ",".join(local_votacion) if local_votacion != [""] else None
                )

                return report.render_to_response(tipo)

            else:
                data["error"] = "No ha ingresado una opción"
        except Exception as e:
            data["error"] = str(e)
        return HttpResponse(json.dumps(data), content_type="application/json")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = ReportForm(usuario=self.request.user)
        context["title"] = "Reporte de Estadisticas Votos Positivos vs Negativos"
        context["action"] = "report"
        return context

    """Electores por Barrios y Manzanas Planilla Visita Casa por Casa"""


class RptElectoral004ReportView(ModuleMixin, FormView):
    template_name = "electoral/reports/rpt_electoral004.html"
    form_class = ReportForm

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        action = request.POST["action"]
        data = {}
        # print(request.POST)
        try:
            if action == "report":
                data = []
                tipo = request.POST["tipo"]
                local_votacion = (
                    request.POST.getlist("local_votacion")
                    if "local_votacion" in request.POST
                    else None
                )
                barrio = (
                    request.POST.getlist("barrio") if "barrio" in request.POST else None
                )
                manzana = (
                    request.POST.getlist("manzana")
                    if "manzana" in request.POST
                    else None
                )
                salto_pagina = (
                    request.POST.getlist("salto_pagina")
                    if "salto_pagina" in request.POST
                    else None
                )
                titulo_extra = (
                    request.POST.getlist("titulo_extra")
                    if "titulo_extra" in request.POST
                    else ""
                )
                filas = (
                    request.POST.getlist("filas") if "filas" in request.POST else None
                )
                # CONFIG
                report = JasperReportBase()
                report.report_name = "rpt_electoral004"

                report.report_url = reverse_lazy(report.report_name)
                report.report_title = (
                    Module.objects.filter(url=report.report_url).first().name
                )
                if len(titulo_extra):
                    report.report_title = titulo_extra[0]
                # PARAMETROS
                report.params["P_LOCAL_VOTACION_ID"] = (
                    ",".join(local_votacion) if local_votacion != [""] else None
                )
                report.params["P_BARRIO_ID"] = (
                    ",".join(barrio) if barrio != [""] else None
                )
                report.params["P_MANZANA_ID"] = (
                    ",".join(manzana) if manzana != [""] else None
                )
                report.params["P_FILAS"] = ",".join(filas) if filas != [""] else None

                if not salto_pagina:
                    report.report_name = "rpt_electoral001_ss"

                return report.render_to_response(tipo)

            else:
                data["error"] = "No ha ingresado una opción"
        except Exception as e:
            data["error"] = str(e)
        return HttpResponse(json.dumps(data), content_type="application/json")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = ReportForm(usuario=self.request.user)
        context["title"] = "Planilla de Electores para Visita Casa x Casa"
        context["action"] = "report"
        return context


# """Planilla de Mesa para Local de Votación - Generación PDF con WeasyPrint"""
class PlanillaMesaPDFView(ModuleMixin, FormView):
    template_name = "electoral/reports/planilla_mesa_filter.html"
    form_class = FormFilterGenerarPDFMesa

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        action = request.POST.get("action")
        data = {}

        try:
            if action == "generate_pdf":
                local_id = request.POST.get("local_votacion")
                mesa_num = request.POST.get("mesa")
                tipo_planilla = request.POST.get("tipo_planilla")

                if not local_id:
                    return HttpResponse(
                        "Error: El ID del Local de Votación es requerido.", status=400
                    )

                local_obj = LocalVotacion.objects.filter(id=local_id).first()
                nombre_local = (
                    local_obj.denominacion if local_obj else f"LOCAL N° {local_id}"
                )

                elector_filter = Elector.objects.filter(local_votacion_id=local_id)

                # --- OPTIMIZACIÓN CRÍTICA CON CAST PARA EL REPORTE COMPLETO ---
                if mesa_num and mesa_num.strip():
                    m_num = int(mesa_num)
                    datos_mesas = (
                        elector_filter.filter(mesa=m_num)
                        .order_by()
                        .values("mesa")
                        .annotate(max_orden=Max(Cast("orden", IntegerField())))
                    )
                    if not datos_mesas:
                        datos_mesas = [{"mesa": m_num, "max_orden": 0}]
                else:
                    # Convertimos el CharField 'mesa' a IntegerField en la query para ordenar rápido
                    datos_mesas = (
                        elector_filter.annotate(mesa_int=Cast("mesa", IntegerField()))
                        .order_by()
                        .values("mesa", "mesa_int")
                        .annotate(max_orden=Max(Cast("orden", IntegerField())))
                        .order_by("mesa_int")
                    )

                mesas_data = []
                columnas = 20

                for d in datos_mesas:
                    m = d["mesa"]

                    try:
                        max_orden = int(d["max_orden"]) if d["max_orden"] else 0
                    except (ValueError, TypeError):
                        max_orden = 0

                    lista_completa = list(range(1, max_orden + 1))
                    matriz_ordenes = [
                        lista_completa[i : i + columnas]
                        for i in range(0, len(lista_completa), columnas)
                    ]

                    if matriz_ordenes:
                        while len(matriz_ordenes[-1]) < columnas:
                            matriz_ordenes[-1].append(None)

                    total_filas = len(matriz_ordenes)
                    if total_filas > 15:
                        escala_css = "escala-micro"
                    elif total_filas > 10:
                        escala_css = "escala-chica"
                    else:
                        escala_css = "escala-normal"

                    mesas_data.append(
                        {
                            "numero_mesa": m,
                            "matriz": matriz_ordenes,
                            "max_orden": max_orden,
                            "escala_css": escala_css,
                        }
                    )

                context = {
                    "nombre_local": nombre_local.upper(),
                    "mesas": mesas_data,
                }

                # 2. SELECCIÓN DINÁMICA DEL TEMPLATE SEGIN EL FILTRO
                if tipo_planilla == "doble":
                    template_reporte = "electoral/reports/planilla_mesa_pdf_doble.html"
                    sufijo_archivo = "_DOBLE"
                else:
                    template_reporte = "electoral/reports/planilla_mesa_pdf.html"
                    sufijo_archivo = ""

                html_string = render_to_string(template_reporte, context)

                if HTML is None:
                    return HttpResponse(
                        "Error: WeasyPrint no está configurado en el servidor.",
                        status=500,
                    )

                html = HTML(
                    string=html_string, base_url=request.build_absolute_uri("/")
                )
                pdf_file = html.write_pdf()

                clean_filename = nombre_local.replace(" ", "_")
                filename = (
                    f"Planilla_{clean_filename}{sufijo_archivo}.pdf"
                    if not mesa_num
                    else f"Planilla_{clean_filename}_Mesa_{mesa_num}.pdf"
                )

                response = HttpResponse(pdf_file, content_type="application/pdf")
                response["Content-Disposition"] = f'inline; filename="{filename}"'
                return response

            else:
                data["error"] = "No ha ingresado una opción válida"
                return HttpResponse(
                    json.dumps(data), content_type="application/json", status=400
                )

        except Exception as e:
            return HttpResponse(f"Error procesando reporte: {str(e)}", status=500)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = FormFilterGenerarPDFMesa(usuario=self.request.user)
        context["title"] = "Planilla de Electores para Visita Casa x Casa"
        context["action"] = "generate_pdf"
        return context
